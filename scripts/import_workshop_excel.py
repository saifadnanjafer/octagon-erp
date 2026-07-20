#!/usr/bin/env python3
import argparse
import copy
import datetime as dt
import json
import math
import shutil
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_EXCEL = Path(__file__).resolve().parents[2] / "COMPANY" / "قاعدة_موحدة_نهائية_نظيفة_حتى_2026_06_30.xlsx"
DEFAULT_DB = Path(__file__).resolve().parents[1] / "database.json"
DEFAULT_SQLITE = Path(__file__).resolve().parents[1] / "database.db"


KNOWN_COLLECTIONS = {
    "employees",
    "contacts",
    "departments",
    "users",
    "locations",
    "quants",
    "stock_moves",
    "transfers",
    "journals",
    "journal_entries",
    "account_moves",
    "account_payments",
    "account_partial_reconciles",
    "payments",
    "maintenance_requests",
    "production_orders",
    "work_orders",
    "audit_log",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return value


def text(value):
    value = clean(value)
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, dt.time):
        return value.strftime("%H:%M")
    return str(value).strip()


def number(value, default=0):
    value = clean(value)
    if value == "":
        return default
    if isinstance(value, str):
        value = value.replace(",", "").replace("،", "").strip()
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def int_number(value, default=0):
    return int(round(number(value, default)))


def normalize_date(value):
    value = clean(value)
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raw = text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def date_display(value):
    date = normalize_date(value)
    return date.strftime("%d/%m/%Y") if date else text(value)


def date_iso(value):
    date = normalize_date(value)
    return date.isoformat() if date else text(value)


def minutes(value):
    value = clean(value)
    if value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, dt.time):
        return value.hour * 60 + value.minute
    raw = text(value)
    if not raw:
        return None
    parts = raw.split(":")
    if len(parts) < 2:
        return None
    try:
        return int(parts[0]) * 60 + int(parts[1])
    except ValueError:
        return None


def sheet_rows(ws, header_row=1):
    headers = [text(v) for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if any(clean(v) != "" for v in row.values()):
            yield row


def load_excel(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    return wb


def find_sheet(wb, startswith):
    for ws in wb.worksheets:
        if ws.title.startswith(startswith):
            return ws
    raise KeyError(f"Sheet starting with {startswith!r} not found")


def build_employee_records(attendance_rows):
    grouped = {}
    for row in attendance_rows:
        name = text(row.get("الموظف"))
        if not name:
            continue
        date = normalize_date(row.get("التاريخ"))
        rec = {
            "day": int_number(row.get("التاريخ") if False else (date.day if date else row.get("اليوم")), 0),
            "month": int_number(row.get("الشهر"), date.month if date else 0),
            "year": int_number(row.get("السنة"), date.year if date else 0),
            "date": date_display(row.get("التاريخ")),
            "weekday": text(row.get("اليوم")),
            "checkIn": text(row.get("دخول")),
            "checkOut": text(row.get("خروج")),
            "checkInMin": minutes(row.get("دخول")),
            "checkOutMin": minutes(row.get("خروج")),
            "hours": number(row.get("ساعات"), 0),
            "status": text(row.get("الحالة")) or "normal",
            "advance": number(row.get("سلفة"), 0),
            "penalty": number(row.get("غرامة"), 0),
            "bonus": number(row.get("مكافأة"), 0),
            "damage": number(row.get("ضرر"), 0),
            "notes": text(row.get("ملاحظات")),
            "fridayAsRegularWork": False,
            "source": "excel_unified_2026_06_30",
        }
        grouped.setdefault(name, []).append(rec)
    for records in grouped.values():
        records.sort(key=lambda r: (r.get("year") or 0, r.get("month") or 0, r.get("day") or 0, r.get("date") or ""))
    return grouped


def build_employees(wb):
    employee_rows = list(sheet_rows(find_sheet(wb, "👥"), 1))
    attendance_rows = list(sheet_rows(find_sheet(wb, "⏱️"), 1))
    records_by_name = build_employee_records(attendance_rows)
    seen_names = {}
    employees = []
    duplicate_names = []
    now = dt.datetime.now().replace(microsecond=0).isoformat()
    for row in employee_rows:
        source_name = text(row.get("الاسم"))
        if not source_name:
            continue
        source_salary = int_number(row.get("الأجر الاسمي"))
        name = source_name
        aliases = []
        if source_name == "عبود" and source_salary == 400000 and "عبدالله هاشم" in records_by_name:
            name = "عبدالله هاشم"
            aliases = ["عبود"]
        seen_names[name] = seen_names.get(name, 0) + 1
        records = records_by_name.get(name, []) if seen_names[name] == 1 else []
        if seen_names[name] > 1:
            duplicate_names.append(name)
        emp = {
            "id": f"EMP_{len(employees) + 1:04d}",
            "empNumber": f"EMP{len(employees) + 1:03d}",
            "name": name,
            "salary": source_salary,
            "prevAdvance": number(row.get("الرصيد السابق"), 0),
            "shift": text(row.get("الشيفت")) or "morning",
            "notes": text(row.get("ملاحظة")),
            "sourceName": source_name,
            "aliases": aliases,
            "sourceRecordCount": int_number(row.get("إجمالي السجلات")),
            "sourceAdvancesTotal": number(row.get("مجموع السلف"), 0),
            "records": records,
            "payments": {},
            "is_active": True,
            "created_at": now,
            "created_by": "excel_import",
            "updated_at": now,
            "updated_by": "excel_import",
            "source": "unified_workshop_excel_2026_06_30",
        }
        if seen_names[name] > 1:
            emp["duplicateNameIndex"] = seen_names[name]
            emp["notes"] = (emp["notes"] + " | " if emp["notes"] else "") + "اسم مكرر في شيت الموظفين؛ لم تُنسخ سجلات الحضور لهذا الصف لتجنب التكرار."
        employees.append(emp)
    return employees, attendance_rows, duplicate_names


def map_tx_type(direction, category, is_advance):
    if text(is_advance) == "نعم" or "سلفة" in text(category):
        return "advance"
    if text(direction) == "داخل":
        return "income"
    return "expense"


CHART_ACCOUNTS = [
    {"id": "cash_workshop", "code": "1001", "name": "قاصة الورشة", "type": "asset", "normal_side": "debit"},
    {"id": "bank_account", "code": "1002", "name": "حساب بنكي / تحويلات", "type": "asset", "normal_side": "debit"},
    {"id": "employee_cash_custody", "code": "1003", "name": "عهد نقدية عند موظفين", "type": "asset", "normal_side": "debit"},
    {"id": "receivables_customers", "code": "1101", "name": "ذمم العملاء", "type": "asset", "normal_side": "debit"},
    {"id": "employee_advances", "code": "1102", "name": "سلف الموظفين", "type": "asset", "normal_side": "debit"},
    {"id": "employee_meal_clearing", "code": "1109", "name": "وسيط توزيع طعام الموظفين", "type": "asset", "normal_side": "debit"},
    {"id": "inventory_stock", "code": "1200", "name": "مخزون مواد", "type": "asset", "normal_side": "debit"},
    {"id": "fixed_assets_tools_machines", "code": "1300", "name": "عدد ومعدات ومكائن", "type": "asset", "normal_side": "debit"},
    {"id": "accumulated_depreciation", "code": "1390", "name": "مجمع إهلاك", "type": "asset", "normal_side": "credit"},
    {"id": "payables_suppliers", "code": "2001", "name": "ذمم الموردين", "type": "liability", "normal_side": "credit"},
    {"id": "customer_deposits", "code": "2002", "name": "دفعات مقدمة من العملاء", "type": "liability", "normal_side": "credit"},
    {"id": "accrued_payroll", "code": "2100", "name": "رواتب مستحقة", "type": "liability", "normal_side": "credit"},
    {"id": "payables_people", "code": "2101", "name": "ذمم الأشخاص/الموظفين", "type": "liability", "normal_side": "credit"},
    {"id": "owner_loans_funding", "code": "2300", "name": "تمويل/قروض من المالك", "type": "liability", "normal_side": "credit"},
    {"id": "owner_capital", "code": "3000", "name": "رأس مال المالك", "type": "equity", "normal_side": "credit"},
    {"id": "owner_drawings", "code": "3100", "name": "مسحوبات المالك", "type": "equity", "normal_side": "debit"},
    {"id": "retained_earnings", "code": "3200", "name": "أرباح/خسائر مرحلة", "type": "equity", "normal_side": "credit"},
    {"id": "opening_balances", "code": "3900", "name": "أرصدة افتتاحية", "type": "equity", "normal_side": "credit"},
    {"id": "income_sales", "code": "4001", "name": "واردات مبيعات/خدمات", "type": "income", "normal_side": "credit"},
    {"id": "income_projects", "code": "4002", "name": "إيرادات مشاريع", "type": "income", "normal_side": "credit"},
    {"id": "other_income", "code": "4900", "name": "إيرادات أخرى", "type": "income", "normal_side": "credit"},
    {"id": "cogs_materials", "code": "5000", "name": "تكلفة المواد المباعة", "type": "expense", "normal_side": "debit"},
    {"id": "expense_payroll", "code": "5101", "name": "رواتب وأجور", "type": "expense", "normal_side": "debit"},
    {"id": "expense_employee_benefits", "code": "5102", "name": "منافع وطعام الموظفين", "type": "expense", "normal_side": "debit"},
    {"id": "expense_materials", "code": "5201", "name": "مواد تشغيل", "type": "expense", "normal_side": "debit"},
    {"id": "expense_tools", "code": "5202", "name": "عدد وصيانة", "type": "expense", "normal_side": "debit"},
    {"id": "rent_expense", "code": "5301", "name": "إيجار", "type": "expense", "normal_side": "debit"},
    {"id": "utilities_expense", "code": "5302", "name": "كهرباء وماء", "type": "expense", "normal_side": "debit"},
    {"id": "transport_fuel_expense", "code": "5303", "name": "نقل ووقود", "type": "expense", "normal_side": "debit"},
    {"id": "marketing_ads_expense", "code": "5401", "name": "تسويق وإعلانات", "type": "expense", "normal_side": "debit"},
    {"id": "expense_general", "code": "5299", "name": "مصروفات عامة", "type": "expense", "normal_side": "debit"},
    {"id": "adjustments_differences", "code": "5900", "name": "فروقات وتسويات", "type": "expense", "normal_side": "debit"},
    {"id": "vat_payable", "code": "2200", "name": "ضريبة القيمة المضافة المستحقة (VAT)", "type": "liability", "normal_side": "credit"},
    {"id": "suspense", "code": "9999", "name": "حساب الاستيداع", "type": "asset", "normal_side": "debit"},
]


FINANCE_CATEGORIES = {
    "expense": [
        {"id": "cat_payroll", "name": "رواتب", "accountId": "expense_payroll"},
        {"id": "cat_employee_advance", "name": "سلف موظفين", "accountId": "employee_advances"},
        {"id": "cat_employee_benefits", "name": "طعام ومنافع", "accountId": "expense_employee_benefits"},
        {"id": "cat_materials", "name": "مواد", "accountId": "expense_materials"},
        {"id": "cat_tools", "name": "صيانة وعدد", "accountId": "expense_tools"},
        {"id": "cat_transport", "name": "نقل وتجهيز", "accountId": "expense_general"},
        {"id": "cat_general", "name": "مصروف عام", "accountId": "expense_general"},
        {"id": "cat_suspense", "name": "تسوية/استيداع", "accountId": "suspense"},
    ],
    "income": [
        {"id": "cat_sales", "name": "وارد مبيعات", "accountId": "income_sales"},
        {"id": "cat_service", "name": "خدمة/تصنيع", "accountId": "income_sales"},
        {"id": "cat_customer_payment", "name": "تسديد عميل", "accountId": "receivables_customers"},
        {"id": "cat_suspense_in", "name": "تسوية واردة", "accountId": "suspense"},
    ],
}


def classify_finance_account(direction_ar, category, is_advance):
    category = text(category)
    is_advance = text(is_advance)
    if is_advance == "نعم" or "سلفة" in category:
        return "employee_advances", "cat_employee_advance", "dept_payroll"
    if text(direction_ar) == "داخل":
        if any(token in category for token in ("رصيد", "فرق", "تسوية", "ملاحظة")):
            return "suspense", "cat_suspense_in", "dept_admin"
        return "income_sales", "cat_sales", "dept_sales"
    if any(token in category for token in ("وجبة", "طعام", "ضيافة", "مشروبات")):
        return "expense_employee_benefits", "cat_employee_benefits", "dept_payroll"
    if any(token in category for token in ("مواد", "شراء", "مورد", "مستلزمات", "طلبية")):
        return "expense_materials", "cat_materials", "dept_workshop"
    if any(token in category for token in ("صيانة", "عدد", "معدات", "تصليح", "مولدة", "أدوات")):
        return "expense_tools", "cat_tools", "dept_workshop"
    if any(token in category for token in ("كروة", "نقل", "بنزين", "مركبات", "سفر", "كهرباء", "ماء", "غاز")):
        return "expense_general", "cat_transport", "dept_workshop"
    if any(token in category for token in ("فرق", "تسوية", "رصيد", "ملاحظة")):
        return "suspense", "cat_suspense", "dept_admin"
    return "expense_general", "cat_general", "dept_workshop"


def build_finance_transactions(wb):
    rows = list(sheet_rows(find_sheet(wb, "📥"), 1))
    txs = []
    for idx, row in enumerate(rows, start=1):
        amount = number(row.get("مبلغ القاصة"), 0)
        if amount == 0:
            continue
        direction_ar = text(row.get("الاتجاه"))
        direction = "in" if direction_ar == "داخل" else "out"
        category = text(row.get("نوع الصرف"))
        is_advance = text(row.get("سلفة موظف؟"))
        tx_type = map_tx_type(direction_ar, category, is_advance)
        account_id, category_id, department_id = classify_finance_account(direction_ar, category, is_advance)
        txs.append({
            "id": f"excel_cashbox_{idx:04d}",
            "date": date_iso(row.get("التاريخ")),
            "createdAt": dt.datetime.now().replace(microsecond=0).isoformat(),
            "type": tx_type,
            "direction": direction,
            "amount": amount,
            "partyName": text(row.get("الجهة")),
            "description": text(row.get("البيان الموحد")) or category,
            "sourceType": "cashbox",
            "importSource": "unified_workshop_excel",
            "sourceId": f"cashbox_row_{text(row.get('ID')) or idx}",
            "departmentId": department_id,
            "categoryId": category_id,
            "accountId": account_id,
            "customerId": "",
            "receiptNo": "",
            "paidByName": "",
            "paymentMethod": "",
            "companyId": "",
            "sourcePeriod": text(row.get("الفترة")),
            "cashboxCategory": category,
            "cashboxEffect": number(row.get("تأثير +/-"), 0),
            "affectsCashbox": text(row.get("يؤثر")),
            "employeeAdvance": is_advance,
            "status": text(row.get("الحالة")),
            "review": text(row.get("مراجعة حسابية")),
        })
    return txs


def build_advances(wb):
    advances = []
    for idx, row in enumerate(sheet_rows(find_sheet(wb, "💵"), 1), start=1):
        amount = number(row.get("المبلغ"), 0)
        advances.append({
            "id": f"advance_excel_{idx:04d}",
            "employeeName": text(row.get("الموظف")),
            "date": date_iso(row.get("التاريخ")),
            "dateDisplay": date_display(row.get("التاريخ")),
            "month": int_number(row.get("الشهر")),
            "year": int_number(row.get("السنة")),
            "amount": amount,
            "description": text(row.get("البيان")),
            "advanceType": text(row.get("النوع")),
            "review": text(row.get("مراجعة")),
            "source": "excel_unified_2026_06_30",
        })
    return advances


def rows_from_offset(wb, sheet_prefix, header_row):
    return list(sheet_rows(find_sheet(wb, sheet_prefix), header_row))


def build_review_payloads(wb):
    attendance_reviews = []
    for row in rows_from_offset(wb, "🔎", 3):
        attendance_reviews.append({
            "id": f"att_review_{int_number(row.get('#'), len(attendance_reviews) + 1):04d}",
            "employeeName": text(row.get("الموظف")),
            "date": date_display(row.get("التاريخ")),
            "weekday": text(row.get("اليوم")),
            "checkIn": text(row.get("دخول")),
            "checkOut": text(row.get("خروج")),
            "hours": number(row.get("ساعات"), 0),
            "status": text(row.get("الحالة")),
            "reviewType": text(row.get("نوع الملاحظة")),
            "details": text(row.get("تفاصيل الملاحظة")),
            "erpStatus": text(row.get("حالة ERP")) or "مفتوح",
            "source": "excel_unified_2026_06_30",
        })

    timesheet_cases = []
    for row in rows_from_offset(wb, "📋", 3):
        timesheet_cases.append({
            "id": f"timesheet_case_{int_number(row.get('#'), len(timesheet_cases) + 1):04d}",
            "employeeName": text(row.get("الموظف")),
            "period": text(row.get("التاريخ / الفترة")),
            "caseType": text(row.get("نوع الحالة")),
            "note": text(row.get("الوضع الحالي / الملاحظة")),
            "proposedAction": text(row.get("الإجراء المقترح بالـERP")),
            "status": text(row.get("الحالة")),
            "source": "excel_unified_2026_06_30",
        })

    account_reviews = []
    for row in rows_from_offset(wb, "🧾", 3):
        account_reviews.append({
            "id": f"account_review_{int_number(row.get('#'), len(account_reviews) + 1):04d}",
            "date": date_iso(row.get("التاريخ")),
            "reviewType": text(row.get("نوع المراجعة")),
            "partyName": text(row.get("الجهة")),
            "amount": number(row.get("المبلغ"), 0),
            "details": text(row.get("التفاصيل")),
            "action": text(row.get("الإجراء")),
            "status": text(row.get("الحالة")),
            "source": "excel_unified_2026_06_30",
        })

    changes = []
    for row in rows_from_offset(wb, "📝", 2):
        changes.append({
            "id": f"change_{int_number(row.get('#'), len(changes) + 1):04d}",
            "type": text(row.get("النوع")),
            "location": text(row.get("الموقع المتأثر")),
            "description": text(row.get("الوصف")),
            "indicator": text(row.get("نوع المؤشر")),
            "source": "excel_unified_2026_06_30",
        })
    return attendance_reviews, timesheet_cases, account_reviews, changes


def build_summary(wb, employees, attendance_rows, finance_transactions, duplicate_names):
    advances = build_advances(wb)
    summary_rows = rows_from_offset(wb, "📊", 2)
    summary = {text(r.get("البند")): {"value": text(r.get("القيمة")), "status": text(r.get("الحالة")), "note": text(r.get("ملاحظة"))} for r in summary_rows if text(r.get("البند"))}
    return {
        "sourceFile": str(DEFAULT_EXCEL),
        "importedAt": dt.datetime.now().replace(microsecond=0).isoformat(),
        "period": "2026-02-15 to 2026-06-30",
        "currency": "IQD",
        "employeeCount": len(employees),
        "attendanceRows": len(attendance_rows),
        "advanceRows": len(advances),
        "advanceTotal": sum(number(row.get("amount"), 0) for row in advances),
        "cashboxTransactions": len(finance_transactions),
        "duplicateEmployeeNames": sorted(set(duplicate_names)),
        "financialSummary": summary,
    }


def ensure_finance_defaults(finance):
    finance.setdefault("cashOpening", 0)
    finance.setdefault("accounts", [])
    finance.setdefault("categories", {"expense": [], "income": []})
    finance.setdefault("departments", [
        {"id": "dept_workshop", "name": "الورشة"},
        {"id": "dept_projects", "name": "المشاريع"},
        {"id": "dept_sales", "name": "المبيعات"},
        {"id": "dept_admin", "name": "الإدارة"},
        {"id": "dept_payroll", "name": "الرواتب"},
    ])
    finance.setdefault("parties", [])
    finance.setdefault("customers", [])
    finance.setdefault("receipts", [])
    existing_accounts = {account.get("id"): account for account in finance["accounts"] if isinstance(account, dict)}
    for account in CHART_ACCOUNTS:
        if account["id"] not in existing_accounts:
            finance["accounts"].append(copy.deepcopy(account))
        else:
            existing_accounts[account["id"]].update({k: v for k, v in account.items() if not existing_accounts[account["id"]].get(k)})
    for side, defaults in FINANCE_CATEGORIES.items():
        finance["categories"].setdefault(side, [])
        existing_categories = {cat.get("id"): cat for cat in finance["categories"][side] if isinstance(cat, dict)}
        for category in defaults:
            if category["id"] not in existing_categories:
                finance["categories"][side].append(copy.deepcopy(category))
            else:
                existing_categories[category["id"]].update({k: v for k, v in category.items() if not existing_categories[category["id"]].get(k)})


def parse_summary_number(summary, key, default=0):
    try:
        value = summary.get(key, {}).get("value", default)
        return number(value, default)
    except Exception:
        return default


def apply_import(db, wb):
    employees, attendance_rows, duplicate_names = build_employees(wb)
    advances = build_advances(wb)
    finance_transactions = build_finance_transactions(wb)
    attendance_reviews, timesheet_cases, account_reviews, changes = build_review_payloads(wb)
    summary = build_summary(wb, employees, attendance_rows, finance_transactions, duplicate_names)

    next_db = copy.deepcopy(db)
    next_db["employees"] = employees
    next_db["selectedEmpIdx"] = 0
    next_db["reportEmpIdx"] = 0

    finance = copy.deepcopy(next_db.get("finance") or {})
    ensure_finance_defaults(finance)
    final_cashbox_balance = parse_summary_number(summary.get("financialSummary", {}), "رصيد القاصة النهائي")
    cashbox_effect = sum(number(tx.get("cashboxEffect"), 0) for tx in finance_transactions)
    finance["cashOpening"] = final_cashbox_balance - cashbox_effect
    finance["transactions"] = finance_transactions
    finance["parties"] = []
    finance["receipts"] = []
    next_db["finance"] = finance

    omni = copy.deepcopy(next_db.get("omni") or {})
    omni["workshopAdvances"] = advances
    omni["employeeAttendance"] = attendance_reviews
    omni["workshopTimesheetCases"] = timesheet_cases
    omni["workshopAccountReviews"] = account_reviews
    omni["workshopExcelChangeLog"] = changes
    omni["workshopExcelImport"] = summary
    omni["historyLedger"] = [{
        "id": "hist_excel_import_2026_06_30",
        "type": "data_import",
        "title": "اعتماد قاعدة Excel الموحدة حتى 2026-06-30",
        "description": f"تم استبدال بيانات الموظفين والحضور والقاصة من ملف Excel. الموظفون: {summary['employeeCount']}، الحضور: {summary['attendanceRows']}، القاصة: {summary['cashboxTransactions']}.",
        "createdAt": summary["importedAt"],
        "source": "excel_unified_2026_06_30",
    }]
    next_db["omni"] = omni

    next_db["audit_log"] = [{
        "id": "audit_excel_import_2026_06_30",
        "action": "workshop_excel_import",
        "createdAt": summary["importedAt"],
        "createdBy": "codex",
        "details": summary,
    }]
    return next_db, summary


def extract_db_collections(obj, path="", collections=None, metadata=None):
    if collections is None:
        collections = {}
    if metadata is None:
        metadata = {}
    if obj is None:
        return collections, metadata

    is_known_collection = path in KNOWN_COLLECTIONS or (path.startswith("omni.") and isinstance(obj, list))
    if isinstance(obj, list):
        has_ids = bool(obj) and all(isinstance(x, dict) and x.get("id") is not None for x in obj)
        if is_known_collection or has_ids:
            collections[path] = obj
            return collections, metadata

    if isinstance(obj, dict) and (path == "" or path in {"omni", "finance"}):
        for key, value in obj.items():
            next_path = f"{path}.{key}" if path else key
            extract_db_collections(value, next_path, collections, metadata)
        return collections, metadata

    metadata[path] = obj
    return collections, metadata


def save_sqlite(sqlite_path, db):
    collections, metadata = extract_db_collections(db)
    con = sqlite3.connect(sqlite_path)
    try:
        con.execute("CREATE TABLE IF NOT EXISTS metadata (key TEXT PRIMARY KEY, value TEXT)")
        con.execute("CREATE TABLE IF NOT EXISTS collections (collection TEXT, id TEXT, data TEXT, PRIMARY KEY (collection, id))")
        con.execute("BEGIN")
        con.execute("DELETE FROM metadata")
        con.execute("DELETE FROM collections")
        for key, value in metadata.items():
            con.execute("INSERT INTO metadata (key, value) VALUES (?, ?)", (key, json.dumps(value, ensure_ascii=False)))
        seen = set()
        for collection, records in collections.items():
            for idx, rec in enumerate(records, start=1):
                raw_id = rec.get("id") if isinstance(rec, dict) else None
                rec_id = str(raw_id or f"{collection[:3]}_{idx:04d}")
                key = (collection, rec_id)
                counter = 1
                while key in seen:
                    rec_id = f"{rec_id}_dup{counter}"
                    key = (collection, rec_id)
                    counter += 1
                seen.add(key)
                if isinstance(rec, dict) and rec.get("id") != rec_id:
                    rec["id"] = rec_id
                con.execute(
                    "INSERT INTO collections (collection, id, data) VALUES (?, ?, ?)",
                    (collection, rec_id, json.dumps(rec, ensure_ascii=False)),
                )
        con.commit()
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()


def backup_file(path, stamp):
    if not path.exists():
        return None
    backup = path.with_name(f"{path.name}.backup-before-excel-import-{stamp}")
    shutil.copy2(path, backup)
    return backup


def main():
    parser = argparse.ArgumentParser(description="Import unified workshop Excel data into Octagon ERP storage.")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--sqlite", type=Path, default=DEFAULT_SQLITE)
    parser.add_argument("--apply", action="store_true", help="Write database.json and database.db. Without this, only prints summary.")
    args = parser.parse_args()

    with args.db.open("r", encoding="utf-8") as fh:
        db = json.load(fh)
    wb = load_excel(args.excel)
    next_db, summary = apply_import(db, wb)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if not args.apply:
        print("DRY RUN ONLY: pass --apply to write changes.")
        return

    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    db_backup = backup_file(args.db, stamp)
    sqlite_backup = backup_file(args.sqlite, stamp)
    args.db.write_text(json.dumps(next_db, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.sqlite.exists():
        save_sqlite(args.sqlite, copy.deepcopy(next_db))
    print(f"WROTE {args.db}")
    if db_backup:
        print(f"BACKUP {db_backup}")
    if args.sqlite.exists():
        print(f"WROTE {args.sqlite}")
    if sqlite_backup:
        print(f"BACKUP {sqlite_backup}")


if __name__ == "__main__":
    main()
